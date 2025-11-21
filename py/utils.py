import os
import re
import sys
import time
import openpyxl
from pathlib import Path
from docxtpl import DocxTemplate


# -------------------- Progress Bar -------------------- #
def progressbar(it, prefix="", size=60, out=sys.stdout):
    count = len(it)
    start = time.time()

    def show(j):
        x = int(size * j / count)
        remaining = ((time.time() - start) / j) * (count - j) if j else 0
        mins, sec = divmod(remaining, 60)
        time_str = f"{int(mins):02}:{sec:03.1f}"
        print(
            f"{prefix}[{'█' * x}{'.' * (size - x)}] {j}/{count} Est wait {time_str}",
            end="\r",
            file=out,
            flush=True,
        )

    if len(it) > 0:
        show(0.1)
        for i, item in enumerate(it):
            yield item
            show(i + 1)
        print(flush=True, file=out)


# -------------------- File Utilities -------------------- #
def safe_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def unique_file_name(path: str) -> str:
    directory = os.path.dirname(path)
    filename, extension = os.path.splitext(os.path.basename(path))
    filename = safe_filename(filename)

    # Remove existing trailing (n)
    base_name = re.sub(r"\s*\(\d+\)$", "", filename)

    counter = 1
    new_path = os.path.join(directory, f"{base_name}{extension}")

    while Path(new_path).is_file():
        new_path = os.path.join(directory, f"{base_name} ({counter}){extension}")
        counter += 1

    return new_path


def load_excel_mapping(
    mapping_path, default_mappings, excel_mappings, sheet_name="File Completion Tool"
):

    mapping_path = Path(mapping_path)
    if not mapping_path.exists():
        print(f"Config file not found: {mapping_path.absolute()}")
        print("Please create 'config.xlsx' in the current directory or visit")
        print(
            "https://github.com/WebDevBernard/Python-Automations to download the template."
        )
        print("\nExiting in ", end="")
        for i in range(10, 0, -1):
            print(f"{i} ", end="", flush=True)
            time.sleep(1)
        print()
        raise FileNotFoundError(f"Config file not found: {mapping_path}")

    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]
    return {key: ws[cell].value for key, cell in excel_mappings.items()}


def write_to_new_docx(
    template_path: Path | None = None, data: dict = None, output_dir: Path | None = None
) -> bool:
    try:
        # Auto-detect template if not provided
        if template_path is None:
            assets_dir = Path.cwd() / "assets"

            if not assets_dir.exists():
                print(
                    "Assets folder not found. Please create an 'assets' folder with the template."
                )
                print("\nExiting in ", end="")
                for i in range(10, 0, -1):
                    print(f"{i} ", end="", flush=True)
                    time.sleep(1)
                print()
                return False

            docx_files = list(assets_dir.glob("*.docx"))

            if not docx_files:
                print(
                    "Template not found. Visit https://github.com/WebDevBernard/Python-Automations to download the template."
                )
                print("\nExiting in ", end="")
                for i in range(10, 0, -1):
                    print(f"{i} ", end="", flush=True)
                    time.sleep(1)
                print()
                return False

            template_path = docx_files[0]
        else:
            template_path = Path(template_path)
            if not template_path.exists():
                print(
                    "Template not found. Visit https://github.com/WebDevBernard/Python-Automations to download the template."
                )
                print("\nExiting in ", end="")
                for i in range(10, 0, -1):
                    print(f"{i} ", end="", flush=True)
                    time.sleep(1)
                print()
                return False

        doc = DocxTemplate(template_path)
        doc.render(data)

        named_insured = (
            str(data.get("named_insured", "Unnamed Client")).rstrip(".:").strip()
        )
        output_dir = output_dir or (Path.home() / "Desktop")
        output_filename = output_dir / f"{named_insured} Renewal Letter.docx"
        doc.save(unique_file_name(output_filename))
        return True

    except Exception as e:
        print(f"Error creating document: {e}")
        print("\nExiting in ", end="")
        for i in range(10, 0, -1):
            print(f"{i} ", end="", flush=True)
            time.sleep(1)
        print()
        return False


def build_index(doc):

    page_index = {}
    text_to_location = []

    for page_num, page in enumerate(doc):
        blocks = []
        for block_idx, block in enumerate(page.get_text("blocks")):
            coords = tuple(block[:4])
            text_lines = block[4].split("\n")
            blocks.append({"words": text_lines, "coords": coords})

            for line_idx, line in enumerate(text_lines):
                text_to_location.append(
                    {
                        "normalized": line.lower(),
                        "page": page_num,
                        "block": block_idx,
                        "line": line_idx,
                        "text": line,
                        "coords": coords,
                    }
                )

        page_index[page_num] = blocks

    return page_index, text_to_location


CONFIG_PATH = Path("../config.xlsx")  # Default path, can be overridden

CONFIG_FIELDS = {
    "task": (2, 1, None),
    "broker_name": (6, 1, None),
    "on_behalf": (8, 1, None),
    "risk_type_1": (12, 1, None),
    "named_insured": (14, 1, None),
    "insurer": (15, 1, None),
    "policy_number": (16, 1, None),
    "effective_date": (17, 1, None),
    "address_line_one": (19, 1, None),
    "address_line_two": (20, 1, None),
    "address_line_three": (21, 1, None),
    "risk_address_1": (23, 1, None),
    "number_of_pdfs": (27, 1, 0),
    "drive_letter": (29, 1, None),
}

PRODUCER_RANGE = (33, 49)
