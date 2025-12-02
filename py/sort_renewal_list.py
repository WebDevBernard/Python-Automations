from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins
from utils import unique_file_name

input_dir = Path.home() / "Downloads"
output_dir = Path.home() / "Desktop"


def sort_renewal_list():
    # Find all Excel files
    xlsx_files = list(Path(input_dir).glob("*.xlsx"))
    xls_files = list(Path(input_dir).glob("*.xls"))
    files = xlsx_files + xls_files

    if not files:
        print(f"No Excel files found in {input_dir}. Place them in Downloads.")
        return

    print(f"Processing: {[f.name for f in files]}")

    # Load all files into dataframes
    dfs = []
    for file in files:
        engine = "xlrd" if file.suffix.lower() == ".xls" else "openpyxl"
        df = pd.read_excel(file, engine=engine)
        dfs.append(df)

    # Combine
    df = pd.concat(dfs, ignore_index=True)

    # Columns needed
    column_list = [
        "policynum",
        "ccode",
        "name",
        "pcode",
        "csrcode",
        "insurer",
        "buscode",
        "renewal",
        "Pulled",
        "D/L",
    ]

    # Reindex columns
    df = df.reindex(columns=column_list)

    # Drop duplicates on policynum
    df = df.drop_duplicates(subset=["policynum"], keep=False)

    # Convert renewal date → sort key MMDD and display key DD-MMM
    df["renewal_sort"] = pd.to_datetime(
        df["renewal"], format="%d-%b-%y", errors="coerce"
    )
    df["renewal_disp"] = df["renewal_sort"].dt.strftime("%d-%b")
    df["renewal"] = df["renewal_sort"].dt.strftime("%m%d")

    # Sorting
    df = df.sort_values(["insurer", "renewal", "name"], ascending=[True, True, True])

    # Replace renewal column with formatted version
    df["renewal"] = df["renewal_disp"]
    df = df.drop(columns=["renewal_disp", "renewal_sort"])

    # Add blank spacer rows between insurer groups
    groups = []
    for insurer, group in df.groupby("insurer"):
        groups.append(group)
        groups.append(pd.DataFrame([[None] * len(df.columns)], columns=df.columns))

    df = pd.concat(groups, ignore_index=True).iloc[:-1]

    # Save to output Excel
    output_path = unique_file_name(output_dir / "renewal_list.xlsx")

    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    df.to_excel(writer, sheet_name="Sheet1", index=False)
    writer.close()

    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Set font + alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal="left")

    # Create table
    end_col = chr(65 + df.shape[1] - 1)
    end_row = df.shape[0] + 1
    ref = f"A1:{end_col}{end_row}"

    table = Table(displayName="Table1", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showRowStripes=True,
        showColumnStripes=False,
    )

    ws.add_table(table)

    # Adjust column widths
    for i, col in enumerate(column_list, 1):
        max_len = max(df[col].astype(str).map(len).max(), len(col))
        if col in ["pcode", "csrcode", "Pulled", "D/L"]:
            ws.column_dimensions[chr(64 + i)].width = 5.0
        elif col == "ccode":
            ws.column_dimensions[chr(64 + i)].width = max_len + 4
        elif col == "policynum":
            ws.column_dimensions[chr(64 + i)].width = max_len + 2.5
        else:
            ws.column_dimensions[chr(64 + i)].width = max_len + 1

    # Add borders for Pulled + D/L
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in ["Pulled", "D/L"]:
        col_idx = column_list.index(col) + 1
        for row in range(1, df.shape[0] + 2):
            ws.cell(row=row, column=col_idx).border = border

    # Page formatting
    ws.print_title_rows = "1:1"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToPage = True
    ws.page_margins = PageMargins(
        top=1.91 / 2.54,
        bottom=1.91 / 2.54,
        left=1.78 / 2.54,
        right=0.64 / 2.54,
    )

    wb.save(output_path)

    print("******** Sort Renewal List ran successfully ********")
    print(f"Output file: {output_path}")
